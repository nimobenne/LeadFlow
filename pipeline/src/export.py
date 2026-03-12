"""
export.py — Export lead data to CSV and XLSX formats.
"""

from __future__ import annotations

import logging
import os
from typing import Any

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column order (follows PRD schema order)
# ---------------------------------------------------------------------------

COLUMN_ORDER = [
    # Identity
    "business_name",
    "city",
    "address",
    "domain",
    "website",
    "phone",
    "yell_listing_url",
    "source_type",
    # Contact
    "personal_email",
    "generic_email",
    "best_email",
    "email_source_url",
    "decision_maker_name",
    "decision_maker_role",
    # Validation
    "mx_valid",
    "mailbox_status",
    "catch_all",
    "role_based",
    "domain_matches_brand",
    "last_verified_at",
    # Site analysis
    "booking_url",
    "booking_platform",
    "whatsapp_present",
    "has_contact_form",
    "has_chat_widget",
    "chat_widget_name",
    "book_now_above_fold",
    "mobile_cta_strength",
    "services_visible",
    "pricing_visible",
    "language_detected",
    "multiple_staff",
    "strong_review_count",
    # Pain points (as pipe-separated string in exports)
    "pain_points",
    # Scores
    "fit_score",
    "confidence_score",
    "priority_score",
    "pricing_fit",
    # Personalization
    "personalization_note",
    "likely_missed_lead_issue",
    "outreach_angle",
    # Pipeline meta
    "stage",
    "site_error",
    "created_at",
    "updated_at",
]

# Header fill colour (light blue)
_HEADER_FILL = PatternFill(
    start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
)
_HEADER_FONT = Font(bold=True, size=11)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def export_to_csv(leads: list[dict], filepath: str) -> str:
    """
    Export list of lead dicts to a CSV file.

    Parameters
    ----------
    leads    : list of lead dicts
    filepath : full path for the output file

    Returns
    -------
    Absolute path of the written file.
    """
    _ensure_dir(filepath)
    df = _leads_to_dataframe(leads)
    df.to_csv(filepath, index=False, encoding="utf-8-sig")
    logger.info("Exported %d leads to CSV: %s", len(leads), filepath)
    return os.path.abspath(filepath)


def export_to_xlsx(leads: list[dict], filepath: str) -> str:
    """
    Export list of lead dicts to an XLSX file with formatted header row.

    Parameters
    ----------
    leads    : list of lead dicts
    filepath : full path for the output file

    Returns
    -------
    Absolute path of the written file.
    """
    _ensure_dir(filepath)
    df = _leads_to_dataframe(leads)

    # Write with openpyxl engine to allow post-write styling
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Leads")
        workbook = writer.book
        worksheet = writer.sheets["Leads"]

        # Style header row
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

        # Auto-size columns
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(col_name)),
                *[len(str(val)) for val in df[col_name].fillna("").astype(str)],
                0,
            )
            # Cap at 60 chars, min at 10
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)

        # Freeze header row
        worksheet.freeze_panes = "A2"

    logger.info("Exported %d leads to XLSX: %s", len(leads), filepath)
    return os.path.abspath(filepath)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _leads_to_dataframe(leads: list[dict]) -> pd.DataFrame:
    """Convert lead dicts to a DataFrame with columns in COLUMN_ORDER."""
    records: list[dict] = []
    for lead in leads:
        row: dict[str, Any] = {}
        for col in COLUMN_ORDER:
            val = lead.get(col, "")
            # Serialise lists as pipe-separated strings
            if isinstance(val, list):
                val = " | ".join(str(v) for v in val)
            elif isinstance(val, dict):
                val = str(val)
            elif val is None:
                val = ""
            row[col] = val
        records.append(row)

    df = pd.DataFrame(records, columns=COLUMN_ORDER)
    return df


def _ensure_dir(filepath: str) -> None:
    """Create parent directories for the output file if they don't exist."""
    parent = os.path.dirname(filepath)
    if parent:
        os.makedirs(parent, exist_ok=True)
